# -*- coding: utf-8 -*-
"""
@Time ： 2022/12/2 16:20
@Auth ： leah_ana
@File ：__init__.py.py
@IDE ：PyCharm
@Motto：ABC(Always Be Coding)
"""
import datetime

from openpyxl import Workbook


def main():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 42
    ws.append([1, 2, 3])
    ws['A2'] = datetime.datetime.now()
    wb.save('sample.xlsx')

if __name__ == '__main__':
    main()